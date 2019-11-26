# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="D:\Git\Python\Demo.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# select demo.xlsx
sheet=wb.active
# set value for cell A1=1
#sheet['A1'] = 1
# set value for cell B2=2
#sheet.cell(row=2, column=2).value = "Trần Thị Bích Thy"
#Loop for row
#Loop for cloum
# save workbook 
wb.save(filepath)
